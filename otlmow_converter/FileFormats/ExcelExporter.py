from pathlib import Path
from typing import List, Iterable

from openpyxl import Workbook
from otlmow_model.OtlmowModel.BaseClasses.OTLObject import OTLObject

from otlmow_converter.AbstractExporter import AbstractExporter
from otlmow_converter.FileFormats.DotnotationTableConverter import DotnotationTableConverter
from otlmow_converter.SettingsManager import load_settings, GlobalVariables

load_settings()

xlsx_settings = GlobalVariables.settings['formats']['xlsx']
xlsx_dotnotation_settings = xlsx_settings['dotnotation']
SEPARATOR = xlsx_dotnotation_settings['separator']
CARDINALITY_SEPARATOR = xlsx_dotnotation_settings['cardinality_separator']
CARDINALITY_INDICATOR = xlsx_dotnotation_settings['cardinality_indicator']
WAARDE_SHORTCUT = xlsx_dotnotation_settings['waarde_shortcut']
LIST_AS_STRING = xlsx_settings['cast_list']
DATETIME_AS_STRING = xlsx_settings['cast_datetime']
ALLOW_NON_OTL_CONFORM_ATTRIBUTES = xlsx_settings['allow_non_otl_conform_attributes']
WARN_FOR_NON_OTL_CONFORM_ATTRIBUTES = xlsx_settings['warn_for_non_otl_conform_attributes']
ABBREVIATE_EXCEL_SHEETTITLES = False

class ExcelExporter(AbstractExporter):
    @classmethod
    def from_objects(cls, sequence_of_objects: Iterable[OTLObject], filepath: Path, **kwargs) -> None:
        cardinality_separator = kwargs.get('cardinality_separator', CARDINALITY_SEPARATOR)
        cardinality_indicator = kwargs.get('cardinality_indicator', CARDINALITY_INDICATOR)
        waarde_shortcut = kwargs.get('waarde_shortcut', WAARDE_SHORTCUT)
        list_as_string = kwargs.get('cast_list', LIST_AS_STRING)
        datetime_as_string = kwargs.get('cast_datetime', DATETIME_AS_STRING)
        allow_non_otl_conform_attributes = kwargs.get('allow_non_otl_conform_attributes',
                                                        ALLOW_NON_OTL_CONFORM_ATTRIBUTES)
        warn_for_non_otl_conform_attributes = kwargs.get('warn_for_non_otl_conform_attributes',
                                                            WARN_FOR_NON_OTL_CONFORM_ATTRIBUTES)
        abbreviate_excel_sheettitles = kwargs.get('abbreviate_excel_sheettitles',
                                                         ABBREVIATE_EXCEL_SHEETTITLES)

        table_dict = DotnotationTableConverter.get_tables_per_type_from_data(
            cardinality_separator=cardinality_separator, cardinality_indicator=cardinality_indicator,
            waarde_shortcut=waarde_shortcut, cast_list=list_as_string, cast_datetime=datetime_as_string,
            allow_non_otl_conform_attributes=allow_non_otl_conform_attributes,
            warn_for_non_otl_conform_attributes=warn_for_non_otl_conform_attributes,
            sequence_of_objects=sequence_of_objects, values_as_string=True)

        wb = Workbook()
        if not sequence_of_objects:
            raise ValueError('There are no asset data to export to Excel')

        created_a_sheet = False
        for class_name in table_dict:
            created_a_sheet = cls._create_sheet_by_name(
                wb=wb, class_name=class_name, table_data=table_dict[class_name],
                abbreviate_excel_sheettitles=abbreviate_excel_sheettitles) or created_a_sheet

        if created_a_sheet:
            del wb['Sheet']

        wb.save(filepath)


    @classmethod
    def _create_sheet_by_name(cls, wb: Workbook, class_name: str, table_data: List[dict],
                              abbreviate_excel_sheettitles: bool = False) -> bool:
        if not table_data:
            return False

        data = DotnotationTableConverter.transform_list_of_dicts_to_2d_sequence(
            list_of_dicts=table_data, empty_string_equals_none=True)

        if abbreviate_excel_sheettitles:
            # abbreviates the class_name so it doesn't exceed the 31 character limit of sheet titles in excel
            split_name = class_name.split("#")
            namespace_name = split_name[0]
            subclass_name = split_name[1]
            class_name = f"{namespace_name[:3]}#{subclass_name}"
            class_name = class_name[:31]

        sheet = wb.create_sheet(class_name)
        for row in data:
            sheet.append(row)

        sheet.sheet_format.defaultColWidth = 20
        return True
